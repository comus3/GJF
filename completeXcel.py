import openpyxl

def write_to_excel(enterprise_name, contact_person, date, mail, excel_file="book.xlsx"):
    try:
        # Load the Excel workbook
        wb = openpyxl.load_workbook(excel_file)

        # Select the first sheet (you can modify this if your file has multiple sheets)
        sheet = wb.active

        # Find the next available row
        next_row = sheet.max_row + 1

        # Write the data to the corresponding cells
        sheet.cell(row=next_row, column=1, value=enterprise_name)
        sheet.cell(row=next_row, column=2, value=contact_person)
        sheet.cell(row=next_row, column=3, value=date)
        sheet.cell(row=next_row, column=4, value=mail)

        # Save the changes to the Excel file
        wb.save(excel_file)
        print(f"COMPLETEXCEL: Data written to {excel_file} successfully!")
        return 0

    except Exception as e:
        print(f"COMPLETEXCEL: An error occurred: {e}")
        return 1
